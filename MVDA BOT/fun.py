import string
import itertools
import pandas as pd
from openpyxl import load_workbook, Workbook


try:
    from robot.libraries.BuiltIn import BuiltIn
    from robot.libraries.BuiltIn import _Misc
    import robot.api.logger as logger
    from robot.api.deco import keyword
    ROBOT = False
except Exception:
    ROBOT = False
    
def get_excel_col(n):
    column=list(
    itertools.chain(
        string.ascii_uppercase, 
        (''.join(pair) for pair in itertools.product(string.ascii_uppercase, repeat=2))
))
    length=len(column)
    for index in range(0,length-1):
        col=column[index]
        if n==col:
            return column[index+1]



def write_verticle(book,sheet,colIndex,dataList):
    df = pd.DataFrame(dataList)
    writer = pd.ExcelWriter(book, engine="openpyxl", mode="a", if_sheet_exists="overlay")
    df.to_excel(writer, sheet_name=sheet, startcol=int(colIndex)-1, index=False, header=False,)
    writer.close()



def write_row(book, sheet, dataList):
    wb = load_workbook(book)
    ws = wb[sheet]
    ws.append(dataList)
    wb.save(filename = book)



def apply_formula_on_matched_sheets(book, sheet, colIndex, formulaValue1, formulaValue2):
    i=1
    formulaValue2 = str(formulaValue2)
    wb = load_workbook(book)
    ws = wb[sheet]
    for i, cellObj in enumerate(ws[colIndex], 1): 
        if i<3:
            pass
        else:
            i=str(i)
            cellObj.value = '='+formulaValue1+'!'+colIndex+i+'*Mini_Stdev!'+colIndex+'$'+formulaValue2+'+Mini_Avg!'+colIndex+'$'+formulaValue2
    wb.save(book)